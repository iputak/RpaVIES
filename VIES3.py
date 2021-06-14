import openpyxl
import os
import json

from selenium import webdriver
from selenium.webdriver.support.ui import Select

def excel():
    ## Excel podatci
    path = 'excel/vatNumbers.xlsx' #Putanja do Excel dokumenta
    wb = openpyxl.load_workbook(path) #Otvaranje datoteke
    #print(wb.sheetnames) #Jeli dohvaća datoteku?
    # Excel listovi
    wsMemberState = wb.get_sheet_by_name('Member State')
    wsRequestMemberState = wb.get_sheet_by_name('Request Member State')
    wsResult = wb.get_sheet_by_name('Results')

    # Broj zapisanih redova u Request Member State listu
    brRedova = wsRequestMemberState.max_row

    # URL, Member State, VAT number iz Excela
    url = wsMemberState['A'+'2'].value
    memberStateCodeImport = wsMemberState['B'+'2'].value
    vatNumberImport = wsMemberState['C'+'2'].value

    # Petlja koja se ponavlja koliko ima podataka, tj. redova u tablici
    for i in range(2, brRedova+1): #Početak s drugog reda zbog zaglavlja
        # Requester Member State, Requester VAT number iz Excela
        requesterMemberStateCodeImport = wsRequestMemberState['A'+str(i)].value
        requesterNumberImport = wsRequestMemberState['B'+str(i)].value
        # Povlačenje podataka s weba
        consultationNumbers = web_login(url, memberStateCodeImport, vatNumberImport, requesterMemberStateCodeImport, requesterNumberImport)
        # Zapis podataka u Excel
        wsResult.append ([requesterNumberImport, requesterMemberStateCodeImport, consultationNumbers])
        # Spremanje tablice
        wb.save(path)

# Ulaz u VIES
def web_login(url, memberStateCodeImport, vatNumberImport, requesterMemberStateCodeImport, requesterNumberImport):
    
    path_to_file = os.getcwd()+'\VIES' #Lokacija projekta + mapa VIES
    
    # Chrome dio vezan uz postavke printanja stranice
    chrome_options = webdriver.ChromeOptions()
    settings = {
       "recentDestinations": [{
            "id": "Save as PDF",
            "origin": "local",
            "account": "",
        }],
        "selectedDestinationId": "Save as PDF",
        "version": 2
    }
    prefs = {'printing.print_preview_sticky_settings.appState': json.dumps(settings),
        "download.default_directory": path_to_file,
        'savefile.default_directory': path_to_file
    }

    chrome_options.add_experimental_option('prefs', prefs)
    chrome_options.add_argument('--kiosk-printing')
    CHROMEDRIVER_PATH = './drivers/chromedriver'
    driver = webdriver.Chrome(chrome_options=chrome_options, executable_path=CHROMEDRIVER_PATH)
    
    #driver.maximize_window()
    #driver.minimize_window()

    driver.get(url)
    
    # Dohvaćanje elemenata
    memberStateCode = Select(driver.find_element_by_name("memberStateCode"))
    vatNumber = driver.find_element_by_name("number")
    requesterMemberStateCode = Select(driver.find_element_by_name("requesterMemberStateCode"))
    requesterNumber = driver.find_element_by_name("requesterNumber")
    verifyButton = driver.find_element_by_name("check")

    # Upis podataka i prijava
    memberStateCode.select_by_value(memberStateCodeImport)
    vatNumber.send_keys(vatNumberImport)
    requesterMemberStateCode.select_by_value(requesterMemberStateCodeImport)
    requesterNumber.send_keys(requesterNumberImport)
    verifyButton.click() #Prijava

    # Dohvaćanje /Consultation Numbers/
    consultationNumbers = driver.find_element_by_css_selector('tbody tr:nth-child(8) td:nth-child(2)').text
    
    if consultationNumbers == "" : # Ako /Consultation Numbers/ ne postoji
        consultationNumbers = "//"
    
    driver.execute_script('document.title="{}";'.format(str(requesterMemberStateCodeImport)+str(requesterNumberImport))); # Spremanje pod prilagođenim imenom        
    driver.execute_script('window.print();')
    
    driver.close() #zatvaranje preglednika
    driver.quit()
    return consultationNumbers

excel()



